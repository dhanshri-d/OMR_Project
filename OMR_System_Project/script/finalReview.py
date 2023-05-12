from imutils.perspective import four_point_transform
from imutils import contours
import numpy as np
import imutils
from openpyxl import Workbook
import cv2
from openpyxl.styles import Font, PatternFill
from PIL import Image


#################################Resize Images and Stak######################
def hconcat_resize(img_list, 
                   interpolation 
                   = cv2.INTER_CUBIC):
      # take minimum hights
    h_min = min(img.shape[0] 
                for img in img_list)
      
    # image resizing 
    im_list_resize = [cv2.resize(img,
                       (int(img.shape[1] * h_min / img.shape[0]),
                        h_min), interpolation
                                 = interpolation) 
                      for img in img_list]
      
    # return final image
    return cv2.hconcat(im_list_resize)


##############################################################################


def grade_exam(image):
    # define the answer key which maps the question number
    # to the correct ansawer
    ANSWER_KEY = {0: 0, 1: 1, 2: 0, 3: 0, 4: 2, 5:3, 6:0, 7:2, 8:4, 9:0} # for 10 question

    correct_answers = {
        1: 2,
        2: 2,
        3: 5,
        4: 1,
        5: 3,
        6: 4,
        7: 1,
        8: 3,
        9: 5,
        10:1,
    }
    student_answers = {}

    # load the image, convert it to grayscale, blur it
    # slightly, then find edges
    #image = image

    #image = cv2.imread(image_path)
    gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
    blurred = cv2.GaussianBlur(gray, (5, 5), 0)
    edged = cv2.Canny(blurred, 75, 200)

    # find contours in the edge map, then initialize
    # the contour that corresponds to the document
    cnts = cv2.findContours(edged.copy(), cv2.RETR_EXTERNAL,
                            cv2.CHAIN_APPROX_SIMPLE)
    cnts = imutils.grab_contours(cnts)
    docCnt = None

    # ensure that at least one contour was found
    if len(cnts) > 0:
        # sort the contours according to their size in
        # descending order
        cnts = sorted(cnts, key=cv2.contourArea, reverse=True)

        # loop over the sorted contours
        for c in cnts:
            # approximate the contour
            peri = cv2.arcLength(c, True)
            approx = cv2.approxPolyDP(c, 0.02 * peri, True)

            # if our approximated contour has four points,
            # then we can assume we have found the paper
            if len(approx) == 4:
                docCnt = approx
                break

    # apply a four point perspective transform to both the
    # original image and grayscale image to obtain a top-down view of the paper
    paper = four_point_transform(image, docCnt.reshape(4, 2))
    warped = four_point_transform(gray, docCnt.reshape(4, 2))

    # apply Otsu's thresholding method to binarize the warped
    # piece of paper
    thresh = cv2.threshold(warped, 0, 255,
                            cv2.THRESH_BINARY_INV | cv2.THRESH_OTSU)[1]

    # find contours in the thresholded image, then initialize
    # the list of contours that correspond to questions
    cnts = cv2.findContours(thresh.copy(), cv2.RETR_EXTERNAL,
                            cv2.CHAIN_APPROX_SIMPLE)
    cnts = imutils.grab_contours(cnts)
    questionCnts = []

    # loop over the contours
    for c in cnts:
        (x, y, w, h) = cv2.boundingRect(c)
        ar = w / float(h)
        if w >= 20 and h >= 20 and ar >= 0.9 and ar <= 1.1:
            questionCnts.append(c)

    # sort the question contours top-to-bottom, then initialize
    # the total number of correct answers
    questionCnts = contours.sort_contours(questionCnts,
    	method="top-to-bottom")[0]
    correct = 0

    # each question has 5 possible answers, to loop over the
    # question in batches of 5
    for (q, i) in enumerate(np.arange(0, len(questionCnts), 5)):
        cnts = contours.sort_contours(questionCnts[i:i + 5])[0]
        bubbled = None
        for (j, c) in enumerate(cnts):
            mask = np.zeros(thresh.shape, dtype="uint8")
            cv2.drawContours(mask, [c], -1, 255, -1)
            mask = cv2.bitwise_and(thresh, thresh, mask=mask)
            total = cv2.countNonZero(mask)
            if bubbled is None or total > bubbled[0]:
                bubbled = (total, j)
                
        color = (0, 0, 255)
        k = ANSWER_KEY[q]
        if k == bubbled[1]:
            color = (0, 255, 0)
            correct += 1
            
        student_answers[q+1] = bubbled[1]+1
        cv2.drawContours(paper, [cnts[k]], -1, color, 3)

    # grab the test taker
    #score = (correct / 5.0)*100
    score = (correct/10)*100
    print("[INFO] score: {:.2f}%".format(score))
    cv2.putText(paper, "{:.2f}%".format(score), (10, 30),
    	cv2.FONT_HERSHEY_SIMPLEX, 0.9, (0, 0, 255), 2)

    #finalOutput = cv2.hconcat([image, paper])
    #cv2.imshow("Original", image)
    #cv2.imshow("Exam", paper)

    finalOutput=hconcat_resize([image, paper]) # function calling
    cv2.imshow('Input-Output_Image', finalOutput)
    #cv2.waitKey(0)
    #cv2.imwrite('final_output.png', finalOutput)
    #cv2.imwrite('final_output.png', paper)
    cv2.waitKey(0)


    wb = Workbook()
    ws = wb.active
    ws.title = "Answers"
    ws.cell(row=1, column=1, value="Question")
    ws.cell(row=1, column=2, value="Student Answer")
    ws.cell(row=1, column=3, value="Correct Answer")


    #red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    for i in student_answers:
        ws.cell(row=i+1, column=1, value=str(i))
        ws.cell(row=i+1, column=2, value=student_answers[i])
        ws.cell(row=i+1, column=3, value=correct_answers[i])

    wb.save("Student_Data.xlsx")
    wb.close()

    return score



#image = cv2.imread("Imges/ans1.jpeg")
#
#grade = grade_exam(image)
#
#
#
import pickle
#
#filename = 'OMR_DETECTION.sav'
#pickle.dump(grade_exam(image),open(filename,'wb'))
#loaded_model = pickle.load(open('OMR_DETECTION.sav','rb'))


#image = load_image('example.png')
#grade = grade_exam(image)

# Serialize the grade to a file
#filename = 'grade.sav'
#pickle.dump(grade, open(filename, 'wb'))

# Load the grade from the file
#loaded_grade = pickle.load(open(filename, 'rb'))
